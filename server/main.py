import os
import openpyxl #type: ignore
import pandas as pd #type: ignore
from flask_cors import CORS #type: ignore
from flask import Flask, request, jsonify #type: ignore
from werkzeug.utils import secure_filename ##type: ignore

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__)
CORS(app, resources={
    r"/*": {
        "origins": ["http://localhost:3000"],
        "methods": ["GET", "POST"],
        "allow_headers": ["Content-Type"]
    }
})
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_single_sheet(wb, sheet_name):
    sheet = wb[sheet_name]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    df = pd.DataFrame(data)
    if df.empty:
        return pd.DataFrame()
    
    header_row_index = 1
    for idx, row in df.iterrows():
        if row.notna().any():
            header_row_index = idx
            break
    
    df.columns = df.iloc[header_row_index]
    df = df.iloc[header_row_index + 1:]
    df = df.reset_index(drop=True)
    
    return df

def get_normalized_day(day_value):
    day_mapping = {
        'MON': 'Monday', 'MONDAY': 'Monday',
        'TUE': 'Tuesday', 'TUESDAY': 'Tuesday',
        'WED': 'Wednesday', 'WEDNESDAY': 'Wednesday',
        'THU': 'Thursday', 'THURSDAY': 'Thursday',
        'FRI': 'Friday', 'FRIDAY': 'Friday',
        'SAT': 'Saturday', 'SATURDAY': 'Saturday'
    }
    return day_mapping.get(day_value.strip().upper())

def check_if_lab(row_idx, faculty_column, current_value):
    prev_value = faculty_column.iloc[row_idx - 1] if row_idx > 0 else None
    next_value = faculty_column.iloc[row_idx + 1] if row_idx < len(faculty_column) - 1 else None
    return (prev_value == current_value) or (next_value == current_value)

def create_schedule_entry(cell_value, time_slot, row_idx, faculty_column):
    try:
        is_lab = check_if_lab(row_idx, faculty_column, cell_value)
        return {
            'subject': cell_value,
            'type': 'Lab' if is_lab else 'Lecture',
            'time_slot': int(time_slot) if pd.notna(time_slot) else 0
        }
    except:
        return None

def create_faculty_schedule_dict(processed_df):
    faculty_master = {}
    faculty_names = []
    
    for col in processed_df.columns[2:]:
        if pd.notna(col) and not str(col).isdigit():
            faculty_names.append(col)
    
    for faculty in faculty_names:
        faculty_master[faculty] = {day: [] for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']}
    
    for faculty in faculty_names:
        faculty_column = processed_df[faculty]
        
        for row_idx in range(len(faculty_column)):
            try:
                cell_value = str(faculty_column.iloc[row_idx]).strip()
                if pd.isna(cell_value) or cell_value == '':
                    continue
                
                day_value = str(processed_df.iloc[row_idx, 0]).strip().upper()
                time_slot = processed_df.iloc[row_idx, 1]
                
                day = get_normalized_day(day_value)
                if not day:
                    continue
                
                schedule_entry = create_schedule_entry(
                    cell_value, 
                    time_slot, 
                    row_idx, 
                    faculty_column
                )
                
                if schedule_entry:
                    faculty_master[faculty][day].append(schedule_entry)
                    
            except IndexError:
                continue
    
    return faculty_master

def parse_subject_info(subject_str):
    try:
        parts = subject_str.strip().split()
        if len(parts) < 2:
            return None
            
        subject_code = parts[0]
        class_info = ''.join(parts[1:])
        
        semester = int(class_info[0])
        division = class_info[1]
        batch = class_info[2] if len(class_info) > 2 else None
        
        return {
            'code': subject_code,
            'semester': semester,
            'division': division,
            'batch': batch
        }
    except:
        return None

def transform_to_required_format(faculty_schedules):
    transformed_data = []
    
    for sheet, faculty_data in faculty_schedules.items():
        for faculty_code, schedule in faculty_data.items():
            for day, slots in schedule.items():
                for slot in slots:
                    try:
                        subject_info = parse_subject_info(slot['subject'])
                        if subject_info:
                            entry = {
                                "subject_code": subject_info['code'],
                                "semester": subject_info['semester'],
                                "division": subject_info['division'],
                                "batch": subject_info['batch'],
                                "is_lab": slot['type'] == 'Lab',
                                "time_slot": slot['time_slot'],
                                "day": day,
                                "faculty_code": faculty_code
                            }
                            transformed_data.append(entry)
                    except:
                        continue
    
    return transformed_data

def process_faculty_data(file_path):
    def process_sheet(wb, sheet_name):
        sheet = wb[sheet_name]
        merged_ranges = sheet.merged_cells.ranges
        
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        
        # Copy data
        for row in sheet.rows:
            for cell in row:
                new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
        
        # Process merged cells
        for merged_range in merged_ranges:
            value = sheet.cell(merged_range.min_row, merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    new_sheet.cell(row=row, column=col, value=value)
        
        df = pd.DataFrame(new_sheet.values)
        df.columns = df.iloc[1]
        df = df.iloc[2:]
        df = df.reset_index(drop=True)
        
        return df
    
    def create_faculty_schedule_dict(processed_df):
        faculty_master = {}
        faculty_names = [col for col in processed_df.columns[2:] if pd.notna(col) and not str(col).isdigit()]
        
        for faculty in faculty_names:
            if faculty not in ['L1', 'L2']:
                faculty_master[faculty] = {
                    'Monday': [], 'Tuesday': [], 'Wednesday': [],
                    'Thursday': [], 'Friday': [], 'Saturday': []
                }
        
        day_mapping = {
            'MON': 'Monday', 'MONDAY': 'Monday',
            'TUES': 'Tuesday', 'TUESDAY': 'Tuesday',
            'WED': 'Wednesday', 'WEDNESDAY': 'Wednesday',
            'THUR': 'Thursday', 'THURSDAY': 'Thursday',
            'FRI': 'Friday', 'FRIDAY': 'Friday',
            'SAT': 'Saturday', 'SATURDAY': 'Saturday'
        }
        
        processed_labs = set()
        
        # Process regular faculty columns
        for faculty in faculty_names:
            if faculty in ['L1', 'L2']:
                continue
                
            faculty_column = processed_df[faculty]
            
            for row_idx in range(len(faculty_column)):
                cell_value = faculty_column.iloc[row_idx]
                day_value = str(processed_df.iloc[row_idx, 0]).strip().upper()
                time_slot = processed_df.iloc[row_idx, 1]
                day = day_mapping.get(day_value)
                
                if pd.notna(cell_value) and cell_value != '' and day is not None:
                    is_lab = False
                    if row_idx > 0 and faculty_column.iloc[row_idx-1] == cell_value:
                        is_lab = True
                        lab_key = f"{faculty}_{day}_{cell_value}_{row_idx-1}"
                        if lab_key in processed_labs:
                            continue
                        processed_labs.add(lab_key)
                    elif row_idx < len(faculty_column)-1 and faculty_column.iloc[row_idx+1] == cell_value:
                        is_lab = True
                        lab_key = f"{faculty}_{day}_{cell_value}_{row_idx}"
                        if lab_key in processed_labs:
                            continue
                        processed_labs.add(lab_key)
                    
                    schedule_entry = {
                        'subject': cell_value,
                        'type': 'Lab' if is_lab else 'Lecture',
                        'time_slot': int(time_slot)
                    }
                    
                    faculty_master[faculty][day].append(schedule_entry)
        
        # Process L1 and L2 columns
        for misc_col in ['L1', 'L2']:
            if misc_col in processed_df.columns:
                misc_column = processed_df[misc_col]
                
                for row_idx in range(len(misc_column)):
                    entry = misc_column.iloc[row_idx]
                    if pd.notna(entry) and isinstance(entry, str) and '(' in entry and ')' in entry:
                        subject = entry[:entry.find('(')].strip()
                        faculty = entry[entry.find('(')+1:entry.find(')')].strip()
                        
                        day_value = str(processed_df.iloc[row_idx, 0]).strip().upper()
                        time_slot = processed_df.iloc[row_idx, 1]
                        day = day_mapping.get(day_value)
                        
                        if day:
                            if faculty not in faculty_master:
                                faculty_master[faculty] = {
                                    'Monday': [], 'Tuesday': [], 'Wednesday': [],
                                    'Thursday': [], 'Friday': [], 'Saturday': []
                                }
                            
                            schedule_entry = {
                                'subject': subject,
                                'type': 'Lab' if subject[-1].isdigit() else 'Lecture',
                                'time_slot': int(time_slot)
                            }
                            
                            faculty_master[faculty][day].append(schedule_entry)
        
        return faculty_master
    
    # Main execution
    wb = openpyxl.load_workbook(file_path)
    all_faculty_schedules = {}
    
    for sheet_name in wb.sheetnames:
        processed_df = process_sheet(wb, sheet_name)
        sheet_schedules = create_faculty_schedule_dict(processed_df)
        
        # Merge schedules from different sheets
        for faculty, schedule in sheet_schedules.items():
            if faculty not in all_faculty_schedules:
                all_faculty_schedules[faculty] = schedule
            else:
                for day in schedule:
                    all_faculty_schedules[faculty][day].extend(schedule[day])
    
    return all_faculty_schedules

def extract_division(class_info):
    divisions = []
    parts = class_info.split('/')
    for part in parts:
        div = ''.join(char for char in part if not char.isdigit())
        divisions.append(div)
    return divisions

def parse_subject_string(subject_str):
    if not subject_str or not isinstance(subject_str, str) or 'TUT' in subject_str:
        return None
        
    parts = subject_str.strip().split()
    if len(parts) < 2:
        return None
        
    subject_code = parts[0]
    class_info = parts[1]
    
    # Extract semester
    semester_digits = ''.join(filter(str.isdigit, class_info[:2]))
    if not semester_digits:
        return None
    semester = int(semester_digits[0])
    
    # Process divisions and batch
    divisions = []
    batch = None
    
    if 'ALL' in class_info:
        divisions = ['ALL']
    else:
        division_parts = class_info[1:].split('/')  # Skip semester number
        
        for part in division_parts:
            # Extract division letter
            div = ''.join(char for char in part if char.isalpha()).upper()
            
            # Extract batch including asterisk if present
            batch_part = ''.join(char for char in part if char.isdigit() or char == '*')
            if batch_part:
                batch = batch_part
                
            divisions.append(div)
    
    return {
        'subject_code': subject_code,
        'semester': semester,
        'divisions': divisions,
        'batch': batch,
        'is_lab': batch is not None
    }

def create_division_tables(faculty_schedules):
    division_tables = {}
    semester_divisions = {}
    
    # First pass: collect all divisions
    for faculty_name, faculty_schedule in faculty_schedules.items():
        for day, sessions in faculty_schedule.items():
            for session in sessions:
                parsed = parse_subject_string(session['subject'])
                if parsed:
                    sem = parsed['semester']
                    if sem not in semester_divisions:
                        semester_divisions[sem] = set()
                    if 'ALL' not in parsed['divisions']:
                        for div in parsed['divisions']:
                            semester_divisions[sem].add(div)
    
    # Second pass: populate tables
    for faculty_name, faculty_schedule in faculty_schedules.items():
        for day, sessions in faculty_schedule.items():
            for session in sessions:
                parsed = parse_subject_string(session['subject'])
                if parsed:
                    sem = parsed['semester']
                    target_divisions = []
                    
                    if 'ALL' in parsed['divisions']:
                        target_divisions = list(semester_divisions[sem])
                    else:
                        target_divisions = parsed['divisions']
                    
                    for div in target_divisions:
                        key = f"{sem}{div}"
                        if key not in division_tables:
                            division_tables[key] = []
                        
                        entry = {
                            'Subject': parsed['subject_code'],
                            'Type': 'Lab' if parsed['is_lab'] else 'Lecture',
                            'Batch': parsed['batch'] if parsed['is_lab'] else '-',
                            'Day': day,
                            'Time_Slot': session['time_slot'],
                            'Faculty': faculty_name
                        }
                        
                        division_tables[key].append(entry)
                        
                        if parsed['is_lab']:
                            next_slot_entry = entry.copy()
                            next_slot_entry['Time_Slot'] = session['time_slot'] + 1
                            division_tables[key].append(next_slot_entry)
    
    # Convert to DataFrames
    for key in division_tables:
        df = pd.DataFrame(division_tables[key])
        df = combine_lab_sessions(df)
        division_tables[key] = df
    
    return division_tables

def combine_lab_sessions(df):
    """Combine consecutive lab sessions into single entries"""
    combined_entries = []
    skip_next = False
    
    for i in range(len(df)):
        if skip_next:
            skip_next = False
            continue
            
        current_row = df.iloc[i]
        
        # If this is a lab and there's a next row
        if i < len(df)-1 and current_row['Type'] == 'Lab':
            next_row = df.iloc[i+1]
            
            # If next row is same lab session
            if (next_row['Type'] == 'Lab' and 
                next_row['Subject'] == current_row['Subject'] and
                next_row['Batch'] == current_row['Batch'] and
                next_row['Faculty'] == current_row['Faculty'] and
                next_row['Time_Slot'] == current_row['Time_Slot'] + 1):
                
                # Combine into single entry
                entry = current_row.to_dict()
                entry['Time_Slot'] = f"{current_row['Time_Slot']}-{next_row['Time_Slot']}"
                combined_entries.append(entry)
                skip_next = True
            else:
                combined_entries.append(current_row.to_dict())
        else:
            combined_entries.append(current_row.to_dict())
    
    new_df = pd.DataFrame(combined_entries)
    return new_df.sort_values(['Day', 'Time_Slot', 'Batch']).reset_index(drop=True)

def create_condensed_tables(division_tables):
    condensed_tables = {}
    
    for div_key, df in division_tables.items():
        # Drop time-related columns
        condensed_df = df.drop(['Time_Slot', 'Day'], axis=1)
        
        # Remove duplicate entries while preserving unique combinations
        condensed_df = condensed_df.drop_duplicates()
        
        # Sort by Subject and Batch for better readability
        condensed_df = condensed_df.sort_values(['Subject', 'Batch']).reset_index(drop=True)
        
        condensed_tables[div_key] = condensed_df
    
    return condensed_tables

def create_final_dictionary(condensed_division_tables, college="LDRP-ITR", department="CE"):
    final_dictionary = {
        college: {
            department: {}
        }
    }
    
    # Group by semesters
    for div_key, df in condensed_division_tables.items():
        semester = div_key[0]  # First character is semester number
        division = div_key[1:] # Rest is division name
        
        # Initialize semester if not exists
        if semester not in final_dictionary["LDRP-ITR"]["CE"]:
            final_dictionary["LDRP-ITR"]["CE"][semester] = {}
            
        # Initialize division
        if division not in final_dictionary["LDRP-ITR"]["CE"][semester]:
            final_dictionary["LDRP-ITR"]["CE"][semester][division] = {}
            
        # Process each subject
        for subject in df['Subject'].unique():
            subject_data = df[df['Subject'] == subject]
            
            subject_dict = {
                'lectures': {},
                'labs': {}
            }
            
            # Process lectures
            lectures = subject_data[subject_data['Type'] == 'Lecture']
            if not lectures.empty:
                subject_dict['lectures'] = {
                    'designated_faculty': lectures['Faculty'].iloc[0]
                }
            
            # Process labs
            labs = subject_data[subject_data['Type'] == 'Lab']
            if not labs.empty:
                subject_dict['labs'] = {
                    batch: {'designated_faculty': faculty}
                    for batch, faculty in zip(labs['Batch'], labs['Faculty'])
                }
            
            final_dictionary["LDRP-ITR"]["CE"][semester][division][subject] = subject_dict
            
    return final_dictionary

def final_func(matrix_file, college="LDRP-ITR", department="CE"):
    faculty_schedules = process_faculty_data(matrix_file)

    division_tables = create_division_tables(faculty_schedules)

    condensed_division_tables = create_condensed_tables(division_tables)
    
    final_dict = create_final_dictionary(condensed_division_tables, college=college, department=department)
    
    return final_dict

@app.route('/process-faculty-data/', methods=['POST'])
def process_faculty_data():
    try:
        if 'file' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'No file uploaded'
            }), 400

        file = request.files['file']
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        file.save(filepath)
        wb = openpyxl.load_workbook(filepath)
        
        processed_data = {}
        for sheet_name in wb.sheetnames:
            processed_df = process_single_sheet(wb, sheet_name)
            faculty_schedule = create_faculty_schedule_dict(processed_df)
            processed_data[sheet_name] = faculty_schedule

        transformed_data = transform_to_required_format(processed_data)
        
        return jsonify({
            'status': 'success',
            'data': transformed_data
        })

    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

    finally:
        if os.path.exists(filepath):
            os.remove(filepath)

@app.route('/process-faculty-matrix/', methods=['POST'])
def process_faculty_matrix():
    try:
        # Validate file presence
        if 'matrix_file' not in request.files or 'abbreviations_file' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'Both matrix file and abbreviations file are required'
            }), 400

        matrix_file = request.files['matrix_file']
        abbreviations_file = request.files['abbreviations_file']

        # Validate file types
        if not (allowed_file(matrix_file.filename) and allowed_file(abbreviations_file.filename)):
            return jsonify({
                'status': 'error',
                'message': 'Invalid file format. Only .xlsx and .xls files are allowed'
            }), 400

        # Create unique filenames
        matrix_filename = secure_filename(f"matrix_{matrix_file.filename}")
        abbrev_filename = secure_filename(f"abbrev_{abbreviations_file.filename}")
        
        matrix_filepath = os.path.join(app.config['UPLOAD_FOLDER'], matrix_filename)
        abbrev_filepath = os.path.join(app.config['UPLOAD_FOLDER'], abbrev_filename)

        # Save files
        matrix_file.save(matrix_filepath)
        abbreviations_file.save(abbrev_filepath)

        # Process files
        faculty_abbreviations = pd.read_excel(abbrev_filepath, sheet_name="Faculty Abbrevations")
        subject_abbreviations = pd.read_excel(abbrev_filepath, sheet_name="Subject Abbrevations")

        # Generate final dictionary
        final_dict = final_func(matrix_filepath, college="LDRP-ITR", department="CE")

        return jsonify({
            'status': 'success',
            'data': final_dict,
            'message': 'Files processed successfully'
        })

    except pd.errors.EmptyDataError:
        return jsonify({
            'status': 'error',
            'message': 'One of the uploaded files is empty'
        }), 400
    
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error processing files: {str(e)}'
        }), 500

    finally:
        # Clean up files
        for filepath in [matrix_filepath, abbrev_filepath]:
            if os.path.exists(filepath):
                os.remove(filepath)

if __name__ == '__main__':
    app.run(port=8000, debug=True)